import streamlit as st
import pandas as pd
import plotly.express as px
import base64
from pathlib import Path
from datetime import datetime, date
import os
import io
import re
from openai import AzureOpenAI
from openpyxl import load_workbook
import pdfplumber


# ======================================================
# CONFIGURAÇÃO DA PÁGINA
# ======================================================
st.set_page_config(
    page_title="Design for Quality | Volkswagen",
    page_icon="🚗",
    layout="wide"
)


# ======================================================
# CONFIGURAÇÃO DO AZURE OPENAI
# ======================================================
AZURE_OPENAI_API_KEY = os.getenv("AZURE_OPENAI_API_KEY")
AZURE_OPENAI_ENDPOINT = os.getenv("AZURE_OPENAI_ENDPOINT")
DEPLOYMENT = os.getenv("AZURE_OPENAI_DEPLOYMENT")

client = None

if AZURE_OPENAI_API_KEY and AZURE_OPENAI_ENDPOINT:
    client = AzureOpenAI(
        api_key=AZURE_OPENAI_API_KEY,
        api_version="2024-02-15-preview",
        azure_endpoint=AZURE_OPENAI_ENDPOINT
    )


# ======================================================
# IDENTIDADE VISUAL VOLKSWAGEN (design tokens)
# ======================================================
VW_LOGO_URL = "https://upload.wikimedia.org/wikipedia/commons/6/6d/Volkswagen_logo_2019.svg"

# Paleta oficial: Navy #001E50 é a cor primária da marca VW.
# Azul elétrico é usado como cor de destaque/ação (CTA), cinza neutro para fundo,
# e uma escala de cinza-texto para hierarquia tipográfica.
VW_CSS_VARS = """
:root {
    --vw-navy: #001E50;
    --vw-navy-dark: #001133;
    --vw-navy-light: #0D4671;
    --vw-blue: #00B0F0;
    --vw-white: #FFFFFF;
    --vw-bg: #F3F5F8;
    --vw-border: #DCE1E8;
    --vw-text: #1A1D22;
    --vw-text-muted: #5B6472;
    --vw-success: #0F9D58;
    --vw-danger: #C8102E;
    --vw-radius: 10px;
    --vw-font: 'Inter', -apple-system, BlinkMacSystemFont, 'Segoe UI', Arial, sans-serif;
}
"""


def inject_css():
    st.markdown(f"""
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800&display=swap');

    {VW_CSS_VARS}

    html,
    body,
    .stApp {{
        font-family: var(--vw-font) !important;
        color: var(--vw-text);
    }}
    h1, h2, h3, h4, h5, h6,
    p,
    span,
    label,
    div[data-testid="stMarkdownContainer"] {{
            color: var(--vw-text) !important;
        }}
        
    .stApp {{
        background: var(--vw-bg);
    }}

    section[data-testid="stSidebar"] {{
        background: var(--vw-navy);
    }}

    section[data-testid="stSidebar"] * {{
        color: var(--vw-white) !important;
    }}

    /* Cabeçalho institucional fixo */
    .vw-topbar {{
        display: flex;
        align-items: center;
        justify-content: space-between;
        background: var(--vw-navy);
        padding: 14px 28px;
        border-radius: var(--vw-radius);
        margin-bottom: 22px;
        box-shadow: 0 2px 10px rgba(0,30,80,0.18);
    }}

    .vw-topbar-left {{
        display: flex;
        align-items: center;
        gap: 14px;
    }}

    .vw-topbar-left img {{
        height: 30px;
        filter: brightness(0) invert(1);
    }}

    .vw-topbar-title {{
        color: var(--vw-white);
        font-weight: 700;
        font-size: 18px;
        letter-spacing: 0.2px;
    }}

    .vw-topbar-sub {{
        color: rgba(255,255,255,0.65);
        font-size: 12px;
        font-weight: 400;
    }}

    .vw-topbar-right {{
        display: flex;
        align-items: center;
        gap: 18px;
        color: rgba(255,255,255,0.85);
        font-size: 13px;
    }}

    .vw-chip {{
        background: rgba(255,255,255,0.10);
        border: 1px solid rgba(255,255,255,0.18);
        padding: 5px 12px;
        border-radius: 999px;
        font-size: 12px;
        color: var(--vw-white);
    }}

    /* Inputs */
    input[type="text"], input[type="password"] {{
        border-radius: 8px !important;
    }}

    /* Botões primários */
    .stButton > button {{
        background: var(--vw-navy);
        color: var(--vw-white);
        border: none;
        border-radius: 8px;
        font-weight: 600;
        padding: 8px 18px;
        transition: background 0.15s ease;
    }}

    .stButton > button:hover {{
        background: var(--vw-blue);
        color: var(--vw-navy-dark);
    }}

    .stDownloadButton > button {{
        background: var(--vw-white);
        color: var(--vw-navy);
        border: 1.5px solid var(--vw-navy);
        border-radius: 8px;
        font-weight: 600;
    }}

    .stDownloadButton > button:hover {{
        background: var(--vw-navy);
        color: var(--vw-white);
    }}

    /* ============================= */
    /* CARDS DE ACESSO RÁPIDO        */
    /* ============================= */
    .vw-card {{
        height: 138px;
        border-radius: var(--vw-radius);
        padding: 18px 20px;
        position: relative;
        background: var(--vw-navy);
        color: var(--vw-white);
        overflow: hidden;
        border: 1px solid rgba(255,255,255,0.06);
        transition: transform 0.15s ease, box-shadow 0.15s ease;
    }}

    .vw-card:hover {{
        transform: translateY(-2px);
        box-shadow: 0 8px 18px rgba(0,30,80,0.22);
    }}

    .vw-card-accent {{
        position: absolute;
        top: 0; left: 0;
        width: 4px; height: 100%;
        background: var(--vw-blue);
    }}

    .vw-card-eyebrow {{
        font-size: 11px;
        text-transform: uppercase;
        letter-spacing: 0.6px;
        color: var(--vw-blue);
        font-weight: 700;
        margin-bottom: 6px;
    }}

    .vw-card-title {{
        font-size: 15px;
        font-weight: 700;
        line-height: 1.3;
    }}

    .vw-card-glyph {{
        position: absolute;
        right: 12px;
        bottom: -6px;
        font-size: 90px;
        font-weight: 800;
        color: rgba(255,255,255,0.07);
        line-height: 1;
    }}

    .vw-card-locked {{
        height: 138px;
        border-radius: var(--vw-radius);
        padding: 18px 20px;
        background: #E9ECF1;
        color: var(--vw-text-muted);
        display: flex;
        align-items: center;
        justify-content: center;
        font-size: 13px;
        border: 1px dashed var(--vw-border);
    }}

    /* Cards de links/ferramentas */
    .vw-link-card {{
        border-radius: var(--vw-radius);
        padding: 16px;
        background: var(--vw-white);
        border: 1px solid var(--vw-border);
        box-shadow: 0 1px 3px rgba(0,0,0,0.04);
    }}

    .vw-link-card-title {{
        font-size: 14px;
        font-weight: 700;
        color: var(--vw-navy);
        margin-bottom: 4px;
    }}

    .vw-link-card-desc {{
        font-size: 12px;
        color: var(--vw-text-muted);
        margin-bottom: 12px;
    }}

    .vw-tag {{
        font-size: 11px;
        font-weight: 600;
        background: var(--vw-navy);
        color: var(--vw-white);
        padding: 3px 10px;
        border-radius: 999px;
        white-space: nowrap;
    }}

    .vw-link-anchor {{
        color: var(--vw-blue) !important;
        font-weight: 600;
        font-size: 12px;
        text-decoration: none;
    }}

    /* Login */
    .vw-login-box {{
        width: 360px;
        margin: 64px auto 0 auto;
        padding: 32px 28px;
        border-radius: 14px;
        background: rgba(255, 255, 255, 0.92);
        box-shadow: 0 10px 40px rgba(0,0,0,0.25);
        text-align: center;
        border: 1px solid rgba(255,255,255,0.4);
    }}

    .vw-login-box img {{
        height: 42px;
        margin-bottom: 14px;
    }}

    .vw-login-title {{
        font-size: 21px;
        font-weight: 800;
        color: var(--vw-navy);
        margin-top: 2px;
    }}

    .vw-login-sub {{
        font-size: 12.5px;
        color: var(--vw-text-muted);
        margin-bottom: 18px;
        letter-spacing: 0.2px;
    }}

    hr {{
        border-color: var(--vw-border) !important;
    }}
    </style>
    """, unsafe_allow_html=True)


# ======================================================
# USUÁRIOS
# ======================================================
USUARIOS = {
    "aannutb": "12345",
    "ufcmart": "12345",
    "vyplfbt": "12345",
    "gibvvr7": "12345",
    "admin": "admin"
}


# ======================================================
# PERMISSÕES DE ACESSO
# ======================================================
PERMISSOES = {
    "aannutb": ["KPM", "GMP21", "STATUS", "ENTREGA VEICULOS QA", "OVERDUE"],
    "admin": ["KPM", "GMP21", "STATUS", "ENTREGA VEICULOS QA", "OVERDUE"],

    "ufcmart": ["ENTREGA VEICULOS QA"],
    "vyplfbt": ["ENTREGA VEICULOS QA"],
    "gibvvr7": ["ENTREGA VEICULOS QA"]
}


# ======================================================
# SESSÃO
# ======================================================
if "logado" not in st.session_state:
    st.session_state.logado = False

if "usuario" not in st.session_state:
    st.session_state.usuario = ""


# ======================================================
# DATA + KW
# ======================================================
def aplicar_background_login():
    img = Path("login_bg.png")

    if img.exists():
        st.markdown(
            f"""
            <style>
            .stApp {{
                background-image:
                    linear-gradient(rgba(0,19,51,0.55), rgba(0,19,51,0.55)),
                    url("data:image/png;base64,{base64.b64encode(img.read_bytes()).decode()}");
                background-size: cover;
                background-position: center;
            }}
            section[data-testid="stSidebar"] {{
                display: none;
            }}
            </style>
            """,
            unsafe_allow_html=True
        )
    else:
        st.markdown(
            """
            <style>
            .stApp {
                background: linear-gradient(135deg, #001E50 0%, #0D4671 60%, #00B0F0 130%);
            }
            section[data-testid="stSidebar"] {
                display: none;
            }
            </style>
            """,
            unsafe_allow_html=True
        )


def data_kw_atual():
    hoje = datetime.now()
    return f"{hoje.strftime('%d/%m/%Y')} · KW {hoje.isocalendar().week}"


def topbar(usuario=""):
    st.markdown(f"""
    <div class="vw-topbar">
        <div class="vw-topbar-left">
            <img src="{VW_LOGO_URL}" />
            <div>
                <div class="vw-topbar-title">Design for Quality</div>
                <div class="vw-topbar-sub">Engenharia de Protótipo · Qualidade VW do Brasil</div>
            </div>
        </div>
        <div class="vw-topbar-right">
            <span class="vw-chip">📅 {data_kw_atual()}</span>
            {f'<span class="vw-chip">👤 {usuario}</span>' if usuario else ''}
        </div>
    </div>
    """, unsafe_allow_html=True)


# ======================================================
# HELPERS EXCEL
# ======================================================
def _to_float_ptbr(value):
    if value is None:
        return None

    if isinstance(value, (int, float)):
        return float(value)

    s = str(value).strip()

    if not s:
        return None

    s = s.replace(".", "").replace(",", ".")

    try:
        return float(s)
    except Exception:
        return None


def extrair_total_coluna_j_openpyxl(uploaded_file, sheet_name):
    wb = load_workbook(io.BytesIO(uploaded_file.getvalue()), data_only=True)
    ws = wb[sheet_name]

    col_j = 10

    for r in range(ws.max_row, 0, -1):
        v = ws.cell(row=r, column=col_j).value
        num = _to_float_ptbr(v)

        if num is not None:
            return num

    return None


def formatar_moeda_br(x):
    if x is None or (isinstance(x, float) and pd.isna(x)):
        return ""

    return f"R$ {x:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


# ======================================================
# HELPERS PDF
# ======================================================
def _to_float_ptbr_num(s):
    if s is None:
        return None

    if isinstance(s, (int, float)):
        return float(s)

    txt = str(s).strip()

    if not txt:
        return None

    txt = txt.replace(" ", "")
    txt = re.sub(r"[^0-9,\.\-]", "", txt)

    if not txt:
        return None

    txt = txt.replace(".", "").replace(",", ".")

    try:
        return float(txt)
    except Exception:
        return None


def extrair_anos_pdf(file_bytes):
    try:
        with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
            if not pdf.pages:
                return []

            text = pdf.pages[0].extract_text() or ""

    except Exception:
        return []

    t = re.sub(r"\s+", " ", text.replace("\n", " ")).strip()
    years = sorted({int(y) for y in re.findall(r"\b20\d{2}\b", t)})

    return years


def extrair_titulo_pdf(file_bytes):
    pattern = re.compile(r"\b\d{3}-VW\d{3}-[A-Z]{3}-CY\d{2}-\d{2}\b")

    try:
        with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
            if not pdf.pages:
                return None

            page = pdf.pages[0]
            text = page.extract_text() or ""
            text = re.sub(r"\s+", " ", text.replace("\n", " ")).strip()

            m = pattern.search(text)

            if m:
                return m.group(0)

            m2 = re.search(
                r"\bNome do ficheiro\b\s+(\d{3}-VW\d{3}-[A-Z]{3}-CY\d{2}-\d{2})\b",
                text
            )

            if m2:
                return m2.group(1)

    except Exception:
        return None

    return None


def extrair_mis12_mis36_por_ano_pdf(file_bytes, ano_alvo):
    """
    Extrai os valores MIS12 e MIS36 do PDF para o ano informado.
    """
    try:
        with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
            if not pdf.pages:
                return {
                    "ano": int(ano_alvo),
                    "MIS12": None,
                    "MIS36": None,
                    "anos_disponiveis": []
                }

            text = pdf.pages[0].extract_text() or ""

    except Exception:
        return {
            "ano": int(ano_alvo),
            "MIS12": None,
            "MIS36": None,
            "anos_disponiveis": []
        }

    t = re.sub(r"\s+", " ", text.replace("\n", " ")).strip()

    header_match = re.search(r"\bHJ\b(.*?\bMIS36\b)", t)

    if header_match:
        header_part = header_match.group(1)
        mis_cols = re.findall(r"\bMIS\d+\b", header_part)
    else:
        mis_cols = re.findall(r"\bMIS\d+\b", t)

    seen = set()
    mis_cols = [m for m in mis_cols if not (m in seen or seen.add(m))]

    years = sorted({int(y) for y in re.findall(r"\b20\d{2}\b", t)})

    ano = int(ano_alvo)

    if ano not in years:
        return {
            "ano": ano,
            "MIS12": None,
            "MIS36": None,
            "anos_disponiveis": years
        }

    pattern = rf"\b{ano}\b(.*?)(?=\b20\d{{2}}\b|\bDifere\b|\bHJ\b\s*Troca\b)"
    m = re.search(pattern, t)

    if not m:
        return {
            "ano": ano,
            "MIS12": None,
            "MIS36": None,
            "anos_disponiveis": years
        }

    block = m.group(1)

    nums = re.findall(r"-?\d+(?:\.\d{3})*,\d+|-?\d+,\d+", block)
    vals = [_to_float_ptbr_num(n) for n in nums]

    # FIX: a versão anterior tentava usar a lista `mis_cols` como se fosse uma função
    # (`mis_cols(vals[i]...)`), o que gerava um TypeError em tempo de execução.
    # O correto é montar o dicionário {nome_da_coluna: valor}.
    mapping = {
        mis_cols[i]: (vals[i] if i < len(vals) else None)
        for i in range(len(mis_cols))
    }

    return {
        "ano": ano,
        "MIS12": mapping.get("MIS12"),
        "MIS36": mapping.get("MIS36"),
        "anos_disponiveis": years
    }


# ======================================================
# AGENDAMENTO
# ======================================================
def aba_agendamento_veiculos():
    st.subheader("🚗 Agendamento de Veículos")

    dia_selecionado = st.date_input(
        "Selecione o dia do agendamento",
        value=date.today(),
        format="DD/MM/YYYY"
    )

    chave_dia = dia_selecionado.strftime("%Y-%m-%d")

    horarios = [
        "08:00 - 08:30", "08:30 - 09:00",
        "09:00 - 09:30", "09:30 - 10:00",
        "10:00 - 10:30", "10:30 - 11:00",
        "11:00 - 11:30", "11:30 - 12:00",
        "12:00 - 12:30", "12:30 - 13:00",
        "13:00 - 13:30", "13:30 - 14:00",
        "14:00 - 14:30", "14:30 - 15:00",
        "15:00 - 15:30", "15:30 - 16:00",
        "16:00 - 16:30", "16:30 - 17:00"
    ]

    if "agenda_veiculos" not in st.session_state:
        st.session_state.agenda_veiculos = {}

    if chave_dia not in st.session_state.agenda_veiculos:
        st.session_state.agenda_veiculos[chave_dia] = {
            h: {"usuario": "", "descricao": "", "salvo": False}
            for h in horarios
        }

    c_h, c_u, c_d, c_s = st.columns([2, 3, 5, 2])

    c_h.markdown("**Horário**")
    c_u.markdown("**Usuário**")
    c_d.markdown("**Descrição**")
    c_s.markdown("**Salvar**")

    st.divider()

    for h in horarios:
        dados = st.session_state.agenda_veiculos[chave_dia][h]

        col1, col2, col3, col4 = st.columns([2, 3, 5, 2])

        col1.write(h)
        col2.write(dados["usuario"] if dados["salvo"] else "")

        dados["descricao"] = col3.text_input(
            "",
            placeholder="Descreva o motivo da utilização do veículo",
            value=dados["descricao"],
            disabled=dados["salvo"],
            key=f"desc_{chave_dia}_{h}",
            label_visibility="collapsed"
        )

        if dados["salvo"]:
            col4.markdown("✅")
        else:
            if col4.button("Salvar", key=f"save_{chave_dia}_{h}"):
                if dados["descricao"].strip():
                    dados["usuario"] = st.session_state.usuario
                    dados["salvo"] = True
                    st.rerun()
                else:
                    st.warning("Preencha a descrição antes de salvar.")

    st.caption("ℹ️ O usuário só é exibido após o salvamento do agendamento.")


# ======================================================
# COPILOTO IA
# ======================================================
def responder_dashboard(pergunta, historico=None):
    if client is None:
        return "⚠️ Azure OpenAI não configurado. Verifique AZURE_OPENAI_API_KEY e AZURE_OPENAI_ENDPOINT."

    if not DEPLOYMENT:
        return "⚠️ DEPLOYMENT não configurado. Defina AZURE_OPENAI_DEPLOYMENT."

    mensagens = [
        {
            "role": "system",
            "content": "Especialista em Qualidade Automotiva VW. Seja objetivo e claro."
        }
    ]

    if historico:
        mensagens.extend(historico)

    mensagens.append({"role": "user", "content": pergunta})

    resp = client.chat.completions.create(
        model=DEPLOYMENT,
        messages=mensagens,
        temperature=0.3,
        max_tokens=300
    )

    return resp.choices[0].message.content


def pagina_input_budget_gmp21():
    st.subheader("🔎 Consulta de Milestone (GMP21)")

    modelo = st.selectbox(
        "Projeto / Plataforma",
        ["Plataformas - Milestone"]
    )

    col1, col2 = st.columns(2)

    with col1:
        ano = st.selectbox("Ano", [2027, 2028, 2029, 2030, 2031])

    with col2:
        mes = st.selectbox("Mês", list(range(1, 13)))

    tipo_milestone = st.selectbox(
        "Tipo de Milestone",
        ["PLATAFORMA", "HUT", "MOTOR"]
    )

    mapa_milestones = {
        "PLATAFORMA": ["PM/PP", "PD/ZV", "PF", "KF", "PLF", "BF", "LF", "VFF", "PVS", "O-S", "SOP", "ME"],
        "HUT": ["PS", "PM/PP", "PD/ZV", "PF", "KF", "PLF", "BF", "LF", "VFF", "PVS", "O-S", "SOP", "ME"],
        "MOTOR": ["KB-A", "PF-A", "AE", "TF", "BF-A", "HSF", "VFF-A", "PVS-A", "O-S A", "SOP-A"]
    }

    milestone = st.selectbox("Milestone", mapa_milestones[tipo_milestone])

    mapa_tempo = {
        "PLATAFORMA": {
            (2027, 3): "PM/PP",
            (2027, 6): "PD/ZV",
            (2027, 9): "PF",
            (2027, 12): "KF",
            (2028, 7): "PLF",
            (2028, 12): "BF",
            (2029, 6): "LF",
            (2030, 1): "VFF",
            (2030, 3): "PVS",
            (2030, 8): "O-S",
            (2031, 1): "SOP",
            (2031, 4): "ME"
        },

        "HUT": {
            (2027, 1): "PS",
            (2027, 6): "PM/PP",
            (2027, 9): "PD/ZV",
            (2027, 11): "PF",
            (2028, 3): "KF",
            (2028, 10): "PLF",
            (2029, 3): "BF",
            (2029, 9): "LF",
            (2030, 1): "VFF",
            (2030, 3): "PVS",
            (2030, 8): "O-S",
            (2031, 1): "SOP",
            (2031, 4): "ME"
        },

        "MOTOR": {
            (2027, 1): "KB-A",
            (2027, 5): "PF-A",
            (2027, 11): "AE",
            (2028, 7): "TF",
            (2029, 2): "BF-A",
            (2029, 9): "HSF",
            (2029, 12): "VFF-A",
            (2030, 2): "PVS-A",
            (2030, 7): "O-S A",
            (2030, 12): "SOP-A"
        }
    }

    milestone_esperado = mapa_tempo.get(tipo_milestone, {}).get((ano, mes))

    if milestone_esperado:
        if milestone_esperado != milestone:
            st.warning(
                f"⚠️ Para {mes}/{ano} o correto é {milestone_esperado} ({tipo_milestone})"
            )
        else:
            st.success("✅ Milestone correto")
    else:
        st.info("ℹ️ Esse mês não possui milestone definido")


def pagina_copiloto_ia():
    st.subheader("🤖 Copiloto IA")

    with st.expander("🔎 Diagnóstico Azure OpenAI", expanded=False):
        st.write("AZURE_OPENAI_ENDPOINT:", "✅ OK" if AZURE_OPENAI_ENDPOINT else "❌ VAZIO")
        st.write("AZURE_OPENAI_API_KEY:", "✅ OK" if AZURE_OPENAI_API_KEY else "❌ VAZIO")
        st.write("AZURE_OPENAI_DEPLOYMENT:", DEPLOYMENT or "❌ VAZIO")

        if st.button("🧪 Testar conexão com Azure OpenAI"):
            try:
                if client is None:
                    st.error("❌ Azure OpenAI não configurado.")
                else:
                    test = client.chat.completions.create(
                        model=DEPLOYMENT,
                        messages=[{"role": "user", "content": "Responda apenas: OK"}],
                        max_tokens=10,
                        temperature=0
                    )

                    st.success("✅ Conectou! Resposta: " + test.choices[0].message.content)

            except Exception as e:
                st.error("❌ Falhou ao chamar Azure OpenAI.")
                st.exception(e)

    if "chat_messages" not in st.session_state:
        st.session_state.chat_messages = [
            {
                "role": "assistant",
                "content": "Olá! Posso ajudar com KPIs, processos e dúvidas do time de Qualidade."
            }
        ]

    for m in st.session_state.chat_messages:
        with st.chat_message(m["role"]):
            st.markdown(m["content"])

    prompt = st.chat_input("Digite sua pergunta...")

    if prompt:
        st.session_state.chat_messages.append({"role": "user", "content": prompt})

        with st.chat_message("user"):
            st.markdown(prompt)

        try:
            with st.chat_message("assistant"):
                with st.spinner("Pensando..."):
                    historico = [
                        x for x in st.session_state.chat_messages
                        if x["role"] in ("user", "assistant")
                    ]

                    answer = responder_dashboard(prompt, historico=historico)
                    st.markdown(answer)

            st.session_state.chat_messages.append({"role": "assistant", "content": answer})

        except Exception as e:
            st.error("❌ Erro ao chamar o Azure OpenAI.")
            st.exception(e)


# ======================================================
# COMPARATIVO EXCEL
# ======================================================
def Comparativo_Custo_Reparo_Prognose():
    st.subheader("🧮 Custo Médio de Reparo")

    st.markdown("""
    **Como funciona:** envie o Excel com as abas dos projetos e o sistema irá capturar
    o **último valor numérico** encontrado na **Coluna J** de cada aba e comparar.

    ⚠️ Se o total em J for fórmula, salve a planilha antes de enviar.
    """)

    arquivo = st.file_uploader("📄 Upload da planilha Excel", type=["xlsx", "xls"])

    if not arquivo:
        st.info("Envie a planilha para gerar o comparativo automático.")
        return

    try:
        wb = load_workbook(io.BytesIO(arquivo.getvalue()), data_only=True)
        abas = wb.sheetnames

    except Exception as e:
        st.error("Não foi possível abrir a planilha. Verifique se o arquivo não está corrompido.")
        st.exception(e)
        return

    sugestao = [s for s in abas if str(s).startswith("203-")]
    default_sel = sugestao[:3] if len(sugestao) >= 3 else abas[:3]

    abas_sel = st.multiselect(
        "Selecione as abas para comparar",
        options=abas,
        default=default_sel
    )

    if len(abas_sel) < 2:
        st.warning("Selecione pelo menos 2 abas para comparar.")
        return

    st.divider()

    resultados = []

    for aba in abas_sel:
        total_j = extrair_total_coluna_j_openpyxl(arquivo, aba)

        resultados.append({
            "Projeto/Aba": aba,
            "Total Coluna J": total_j
        })

    df = pd.DataFrame(resultados)

    if df["Total Coluna J"].notna().any():
        maxv = df["Total Coluna J"].max()
        minv = df["Total Coluna J"].min()

        df["Diferença p/ Máx"] = maxv - df["Total Coluna J"]
        df["Diferença p/ Mín"] = df["Total Coluna J"] - minv

    col1, col2, col3 = st.columns(3)

    col1.metric(
        "Maior Total (J)",
        formatar_moeda_br(df["Total Coluna J"].max() if df["Total Coluna J"].notna().any() else None)
    )

    col2.metric(
        "Menor Total (J)",
        formatar_moeda_br(df["Total Coluna J"].min() if df["Total Coluna J"].notna().any() else None)
    )

    col3.metric(
        "Delta (Máx - Mín)",
        formatar_moeda_br(
            (df["Total Coluna J"].max() - df["Total Coluna J"].min())
            if df["Total Coluna J"].notna().any()
            else None
        )
    )

    st.markdown("### ✅ Comparativo Custo de Reparo")

    df_view = df.copy()
    df_view["Total Coluna J"] = df_view["Total Coluna J"].apply(formatar_moeda_br)

    if "Diferença p/ Máx" in df_view.columns:
        df_view["Diferença p/ Máx"] = df["Diferença p/ Máx"].apply(formatar_moeda_br)
        df_view["Diferença p/ Mín"] = df["Diferença p/ Mín"].apply(formatar_moeda_br)

    st.dataframe(df_view, use_container_width=True)

    st.divider()
    st.markdown("### 📊 Visual")

    df_plot = df.dropna(subset=["Total Coluna J"]).copy()

    if len(df_plot):
        fig = px.bar(
            df_plot,
            x="Projeto/Aba",
            y="Total Coluna J",
            text="Total Coluna J",
            color_discrete_sequence=["#001E50"]
        )

        fig.update_traces(
            texttemplate="%{text:.2f}",
            textposition="outside"
        )

        fig.update_layout(
            barmode="group",
            title="Comparativo Custo Médio de Reparo",
            uniformtext_minsize=8,
            uniformtext_mode="hide",
            plot_bgcolor="#FFFFFF",
            paper_bgcolor="#FFFFFF",
            font=dict(color="#1A1D22", family="Inter, sans-serif"),
            hovermode="x unified"
        )

        st.plotly_chart(fig, use_container_width=True)

    else:
        st.warning("Não foi possível localizar valores numéricos na Coluna J das abas selecionadas.")

    st.divider()

    st.download_button(
        "Baixar comparativo (CSV)",
        data=df.to_csv(index=False).encode("utf-8"),
        file_name="comparativo_total_coluna_J.csv",
        mime="text/csv"
    )


# ======================================================
# COMPARATIVO PDF
# ======================================================
def Comparativo_MIS_PDF():
    st.subheader("📄 Comparativo MIS12 e MIS36 (PDF)")

    st.markdown("""
    **Como funciona:** envie **2 PDFs**. O sistema identifica o título/código do relatório
    e extrai os valores **MIS12** e **MIS36** do ano selecionado.
    """)

    colA, colB = st.columns(2)

    with colA:
        pdf_a = st.file_uploader("Upload PDF A", type=["pdf"], key="pdf_a")

    with colB:
        pdf_b = st.file_uploader("Upload PDF B", type=["pdf"], key="pdf_b")

    if not pdf_a or not pdf_b:
        st.info("Envie os dois PDFs para iniciar o comparativo.")
        return

    bytes_a = pdf_a.getvalue()
    bytes_b = pdf_b.getvalue()

    with st.spinner("Lendo títulos/códigos dos PDFs..."):
        titulo_a = extrair_titulo_pdf(bytes_a) or "Não identificado"
        titulo_b = extrair_titulo_pdf(bytes_b) or "Não identificado"

    st.markdown("### 🏷️ Identificação dos PDFs")

    c1, c2 = st.columns(2)

    c1.info(f"**PDF A:** {titulo_a}")
    c2.info(f"**PDF B:** {titulo_b}")

    anos_disp = sorted(set(extrair_anos_pdf(bytes_a)) | set(extrair_anos_pdf(bytes_b)))

    if not anos_disp:
        st.error("Não foi possível identificar anos no PDF.")
        return

    default_ano = max(anos_disp)

    ano_sel = st.selectbox(
        "Ano para comparação",
        options=anos_disp,
        index=anos_disp.index(default_ano)
    )

    with st.spinner("Extraindo MIS12/MIS36 do PDF A..."):
        res_a = extrair_mis12_mis36_por_ano_pdf(bytes_a, int(ano_sel))

    with st.spinner("Extraindo MIS12/MIS36 do PDF B..."):
        res_b = extrair_mis12_mis36_por_ano_pdf(bytes_b, int(ano_sel))

    def fmt_num(x):
        if x is None:
            return "—"

        return f"{x:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")

    def delta(a, b):
        if a is None or b is None:
            return None

        return b - a

    def delta_pct(a, b):
        if a is None or b is None or a == 0:
            return None

        return (b - a) / a * 100.0

    d_mis12 = delta(res_a["MIS12"], res_b["MIS12"])
    p_mis12 = delta_pct(res_a["MIS12"], res_b["MIS12"])

    d_mis36 = delta(res_a["MIS36"], res_b["MIS36"])
    p_mis36 = delta_pct(res_a["MIS36"], res_b["MIS36"])

    if (
        res_a["MIS12"] is None
        or res_a["MIS36"] is None
        or res_b["MIS12"] is None
        or res_b["MIS36"] is None
    ):
        st.warning("Algum valor não foi encontrado para este ano em um dos PDFs.")
        st.write("Anos disponíveis PDF A:", res_a.get("anos_disponiveis", []))
        st.write("Anos disponíveis PDF B:", res_b.get("anos_disponiveis", []))

    st.divider()
    st.markdown("### ✅ Métricas")

    c1, c2, c3, c4 = st.columns(4)

    c1.metric("MIS12 - PDF A", fmt_num(res_a["MIS12"]))
    c2.metric("MIS12 - PDF B", fmt_num(res_b["MIS12"]))
    c3.metric("Δ MIS12 (B - A)", fmt_num(d_mis12))
    c4.metric("Δ% MIS12", "—" if p_mis12 is None else f"{p_mis12:.1f}%")

    c1, c2, c3, c4 = st.columns(4)

    c1.metric("MIS36 - PDF A", fmt_num(res_a["MIS36"]))
    c2.metric("MIS36 - PDF B", fmt_num(res_b["MIS36"]))
    c3.metric("Δ MIS36 (B - A)", fmt_num(d_mis36))
    c4.metric("Δ% MIS36", "—" if p_mis36 is None else f"{p_mis36:.1f}%")

    st.divider()
    st.markdown("### 📋 Tabela")

    df = pd.DataFrame([
        {
            "PDF": "A",
            "Título": titulo_a,
            "Ano": int(ano_sel),
            "MIS12": res_a["MIS12"],
            "MIS36": res_a["MIS36"]
        },
        {
            "PDF": "B",
            "Título": titulo_b,
            "Ano": int(ano_sel),
            "MIS12": res_b["MIS12"],
            "MIS36": res_b["MIS36"]
        }
    ])

    df_view = df.copy()
    df_view["MIS12"] = df_view["MIS12"].apply(fmt_num)
    df_view["MIS36"] = df_view["MIS36"].apply(fmt_num)

    st.dataframe(df_view, use_container_width=True)

    st.markdown("### 📊 Visual")

    df_plot = df.dropna(subset=["MIS12", "MIS36"], how="all").copy()

    if len(df_plot):
        df_melt = df_plot.melt(
            id_vars=["PDF", "Título", "Ano"],
            value_vars=["MIS12", "MIS36"],
            var_name="MIS",
            value_name="Valor"
        )

        fig = px.bar(
            df_melt,
            x="MIS",
            y="Valor",
            color="PDF",
            barmode="group",
            text="Valor",
            hover_data=["Título"],
            color_discrete_sequence=["#001E50", "#00B0F0"]
        )

        fig.update_traces(
            texttemplate="%{text:.2f}",
            textposition="outside"
        )

        fig.update_layout(
            plot_bgcolor="#FFFFFF",
            paper_bgcolor="#FFFFFF",
            font=dict(color="#1A1D22", family="Inter, sans-serif")
        )

        st.plotly_chart(fig, use_container_width=True)

    else:
        st.info("Sem valores suficientes para plotar.")

    st.divider()

    st.download_button(
        "Baixar comparativo (CSV)",
        data=df.to_csv(index=False).encode("utf-8"),
        file_name=f"comparativo_MIS12_MIS36_{ano_sel}.csv",
        mime="text/csv"
    )


# ======================================================
# LOGIN
# ======================================================
def tela_login():
    st.markdown(f"""
    <div class="vw-login-box">
        <img src="{VW_LOGO_URL}" />
        <div class="vw-login-title">Design for Quality</div>
        <div class="vw-login-sub">Sistema de Qualidade · Volkswagen do Brasil</div>
    </div>
    """, unsafe_allow_html=True)

    col1, col2, col3 = st.columns([3, 4, 3])

    with col2:
        with st.form("login"):
            user = st.text_input("Usuário REDE VW")
            pwd = st.text_input("Senha", type="password")

            entrar = st.form_submit_button("Entrar", use_container_width=True)

            if entrar:
                user = user.lower().strip()

                if user in USUARIOS and USUARIOS[user] == pwd:
                    st.session_state.logado = True
                    st.session_state.usuario = user
                    st.rerun()
                else:
                    st.error("Usuário ou senha inválidos")


# ======================================================
# LINKS E FERRAMENTAS
# ======================================================
def pagina_links_ferramentas():
    st.subheader("🔗 Links e Ferramentas do Dia a Dia")

    recursos = [
        {
            "nome": "Power BI - KPMs",
            "url": "https://app.powerbi.com/reportEmbed?reportId=e373352c-48e7-4ea0-936e-db63d70c84b1",
            "desc": "Dashboard principal de KPIs",
            "tag": "BI"
        },
        {
            "nome": "SharePoint - Qualidade",
            "url": "https://volkswagengroup.sharepoint.com/:f:/r/sites/QAProttipos/Shared%20Documents/General?csf=1&web=1&e=eNV37D",
            "desc": "Documentos e procedimentos do time",
            "tag": "Docs"
        },
        {
            "nome": "Teams - Squad QA",
            "url": "https://SEU_LINK_AQUI",
            "desc": "Canal do time / comunicação",
            "tag": "Teams"
        },
        {
            "nome": "Pasta de Trabalho - Rede",
            "url": r"G:\ANCBQD01\S1004_B-QP_ Plan_Central Novos_Projetos\S2043_B-QP_VSC_QA_&_Eng_Prototipo\DESIGN FOR QUALITY",
            "desc": "Atalho para pasta da rede",
            "tag": "Files"
        }
    ]

    colA, colB, colC = st.columns(3)

    cols = [colA, colB, colC]

    for i, r in enumerate(recursos):
        with cols[i % 3]:
            url = r["url"]
            is_http = url.lower().startswith("http")

            if is_http:
                link_html = f'<a class="vw-link-anchor" href="{url}" target="_blank">Abrir ↗</a>'
            else:
                link_html = f'<span style="font-size:12px;opacity:.9;">{url}</span>'

            st.markdown(
                f"""
                <div class="vw-link-card">
                  <div class="vw-link-card-title">{r['nome']}</div>
                  <div class="vw-link-card-desc">{r['desc']}</div>
                  <div style="display:flex;justify-content:space-between;align-items:center;gap:10px;">
                    <span class="vw-tag">{r['tag']}</span>
                    <div style="text-align:right;overflow:hidden;text-overflow:ellipsis;white-space:nowrap;max-width:220px;">
                      {link_html}
                    </div>
                  </div>
                </div>
                """,
                unsafe_allow_html=True
            )

            if not is_http:
                st.text_input(
                    "Copiar caminho:",
                    value=url,
                    label_visibility="collapsed",
                    key=f"path_{i}"
                )


# ======================================================
# TEMPLATES
# ======================================================
def pagina_templates():
    st.subheader("📄 Templates e Arquivos")

    pasta = Path("templates")
    pasta.mkdir(exist_ok=True)

    arquivos = sorted(pasta.glob("*"))

    if not arquivos:
        st.info("Nenhum arquivo em /templates ainda. Coloque aqui os modelos.")
        return

    for arq in arquivos:
        col1, col2 = st.columns([7, 2])

        col1.write(f"📌 {arq.name}")

        with col2:
            st.download_button(
                "Baixar",
                data=arq.read_bytes(),
                file_name=arq.name,
                mime="application/octet-stream",
                key=f"dl_{arq.name}"
            )


# ======================================================
# HELPERS DE NAVEGAÇÃO
# ======================================================
def sync_pagina_com_url():
    if "pagina_atual" not in st.session_state:
        if "pagina" in st.query_params:
            pagina_url = st.query_params["pagina"]

            if isinstance(pagina_url, list):
                pagina_url = pagina_url[0]

            st.session_state.pagina_atual = pagina_url

        else:
            st.session_state.pagina_atual = "HOME"


def ir_para(pagina):
    st.session_state.pagina_atual = pagina
    st.query_params.update({"pagina": pagina})
    st.rerun()


def botao_voltar():
    if st.button("⬅️ Voltar", key="btn_voltar"):
        ir_para("HOME")


# ======================================================
# CARD DE ACESSO RÁPIDO (componente reutilizável)
# ======================================================
def render_card(titulo, categoria, glifo, key, tem_acesso):
    if tem_acesso:
        st.markdown(f"""
        <div class="vw-card">
            <div class="vw-card-accent"></div>
            <div class="vw-card-eyebrow">{categoria}</div>
            <div class="vw-card-title">{titulo}</div>
            <div class="vw-card-glyph">{glifo}</div>
        </div>
        """, unsafe_allow_html=True)

        return st.button("Acessar →", key=key, use_container_width=True)
    else:
        st.markdown("""
        <div class="vw-card-locked">🔒 Sem acesso</div>
        """, unsafe_allow_html=True)

        return False


# ======================================================
# PAINEL PRINCIPAL
# ======================================================
def painel():
    sync_pagina_com_url()
    inject_css()

    usuario = st.session_state.get("usuario", "")
    permissoes = PERMISSOES.get(usuario, [])

    topbar(usuario)

    with st.sidebar:
        st.markdown("### Navegação")

        if st.button("🏠 Início", use_container_width=True):
            ir_para("HOME")

        st.markdown("---")
        st.markdown(f"**Usuário:** {usuario}")

        if st.button("🚪 Sair", use_container_width=True):
            logout()

    pagina = st.session_state.get("pagina_atual", "HOME")

    # ======================
    # HOME
    # ======================
    if pagina == "HOME":
        st.markdown("""
        <h2 style="text-align:center; color:#001E50; margin-bottom:2px;">Acesso Rápido</h2>
        <p style="text-align:center; color:#5B6472; font-size:13px; margin-top:0;">
            Selecione um módulo para continuar
        </p>
        """, unsafe_allow_html=True)

        st.write("")

        col1, col2, col3, col4, col5 = st.columns(5)

        with col1:
            if render_card("Overdue Streifenlist", "Qualidade", "O", "overdue", "OVERDUE" in permissoes):
                ir_para("OVERDUE")

        with col2:
            if render_card("KPI KPM", "Indicadores", "K", "kpm", "KPM" in permissoes):
                ir_para("KPM")

        with col3:
            if render_card("Prognose GMP21", "Budget", "G", "gmp21", "GMP21" in permissoes):
                ir_para("GMP21")

        with col4:
            if render_card("Análise Custo Reparo", "Financeiro", "S", "status", "STATUS" in permissoes):
                ir_para("STATUS")

        with col5:
            if render_card("Entrega Veículos", "Logística", "E", "entrega_veiculos_qa", "ENTREGA VEICULOS QA" in permissoes):
                ir_para("ENTREGA VEICULOS QA")

        st.write("")
        st.divider()
        pagina_links_ferramentas()

    # ======================
    # ENTREGA VEICULOS QA
    # ======================
    elif pagina == "ENTREGA VEICULOS QA":
        if "ENTREGA VEICULOS QA" not in permissoes:
            st.warning("🚫 Acesso negado")
            return

        botao_voltar()

        st.markdown("""
                2<h2 style="color:#001E50;">
                🚗 Status Liberações ZP8
                </h2>
                """, unsafe_allow_html=True)
                
        st.caption("Rodagem 2026")
        [data-testid="stCaptionContainer"] {
            color: #5B6472 !important;
            }

        import plotly.graph_objects as go

        try:
            df = pd.read_csv("dados_rodagem.csv")

        except FileNotFoundError:
            st.error("Arquivo dados_rodagem.csv não encontrado.")
            return

        df.columns = df.columns.str.strip()
        df.columns = ["Mes", "Prevista", "Liberados"]

        meses = df["Mes"].tolist()
        prevista = [int(v) if pd.notna(v) else None for v in df["Prevista"]]
        liberados = [int(v) if pd.notna(v) else None for v in df["Liberados"]]

        fig = go.Figure()

        fig.add_trace(go.Bar(
            name="Rodagem Prevista",
            x=meses,
            y=prevista,
            text=[v if v is not None else "" for v in prevista],
            textposition="outside",
            marker_color="#0D4671"
        ))

        fig.add_trace(go.Bar(
            name="Veículos Liberados",
            x=meses,
            y=liberados,
            text=[v if v is not None else "" for v in liberados],
            textposition="outside",
            marker_color="#00B0F0"
        ))

        fig.update_layout(
            margin=dict(b=80),
            barmode="group",
            plot_bgcolor="#FFFFFF",
            paper_bgcolor="#FFFFFF",
            font=dict(color="#1A1D22", family="Inter, sans-serif"),
            hovermode="x unified",
            legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1),
            annotations=[
                dict(
                    text="<b>2026</b>",
                    x=0.5,
                    y=-0.22,
                    xref="paper",
                    yref="paper",
                    showarrow=False,
                    font=dict(size=16, color="#001E50")
                )
            ]
        )

        col1, col2 = st.columns([3, 1])

        with col1:
            st.plotly_chart(fig, use_container_width=True)

        with col2:
            total_prevista = sum(v for v in prevista if v is not None)
            total_liberados = sum(v for v in liberados if v is not None)

            percentual = (
                (total_liberados / total_prevista) * 100
                if total_prevista > 0
                else 0
            )

            st.markdown("#### 📊 Totais")

            st.metric("🚗 Prevista", total_prevista)
            st.metric("✅ Liberados", total_liberados)
            st.metric("🎯 % Liberação", f"{percentual:.1f}%")

            div[data-testid="stMetric"] {
                color: #1A1D22 !important;
                }
                div[data-testid="stMetricValue"] {
                    color: #001E50 !important;
                        font-weight: 700 !important;
                        }
                        div[data-testid="stMetricLabel"] {
                            color: #5B6472 !important;
                            }

    # ======================
    # GMP21
    # ======================
    elif pagina == "GMP21":
        if "GMP21" not in permissoes:
            st.warning("🚫 Acesso negado")
            return

        botao_voltar()
        st.subheader("GMP21 Budget")
        pagina_input_budget_gmp21()

    # ======================
    # KPM
    # ======================
    elif pagina == "KPM":
        if "KPM" not in permissoes:
            st.warning("🚫 Acesso negado")
            return

        botao_voltar()
        st.subheader("Dashboard KPM")

    # ======================
    # STATUS
    # ======================
    elif pagina == "STATUS":
        if "STATUS" not in permissoes:
            st.warning("🚫 Acesso negado")
            return

        botao_voltar()
        Comparativo_Custo_Reparo_Prognose()
        st.divider()
        Comparativo_MIS_PDF()

    # ======================
    # OVERDUE
    # ======================
    elif pagina == "OVERDUE":
        if "OVERDUE" not in permissoes:
            st.warning("🚫 Acesso negado")
            return

        botao_voltar()
        st.subheader("Overdue Streifenlist Dashboard")


# ======================================================
# LOGOUT
# ======================================================
def logout():
    st.session_state.logado = False
    st.session_state.pagina_atual = "HOME"
    st.rerun()


# ======================================================
# FLUXO PRINCIPAL DO APP
# ======================================================
if st.session_state.get("logado", False):
    painel()
else:
    inject_css()
    aplicar_background_login()
    tela_login()
